# Form implementation generated from reading ui file 'ImportTAG_Window.ui'
#
# Created by: PyQt6 UI code generator 6.4.1
#
# WARNING: Any manual changes made to this file will be lost when pyuic6 is
# run again.  Do not edit this file unless you know what you are doing.

import sys
from PyQt6 import QtCore, QtGui, QtWidgets
import pandas as pd
from tkinter.filedialog import askopenfilename
import psycopg2
from config import config
import math
import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

basedir = r"\\nas01\DATOS\Comunes\EIPSA-ERP"


class Ui_ImportTAG_Window(object):
    def setupUi(self, ImportTAG_Window):
        ImportTAG_Window.setObjectName("ImportTAG_Window")
        ImportTAG_Window.resize(640, 330)
        ImportTAG_Window.setMinimumSize(QtCore.QSize(640, 330))
        ImportTAG_Window.setMaximumSize(QtCore.QSize(640, 330))
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        ImportTAG_Window.setWindowIcon(icon)
        ImportTAG_Window.setStyleSheet("QWidget {\n"
"background-color: rgb(255, 255, 255);\n"
"}\n"
"\n"
".QFrame {\n"
"    border: 2px solid black;\n"
"}\n"
"\n"
"QPushButton {\n"
"background-color: #33bdef;\n"
"  border: 1px solid transparent;\n"
"  border-radius: 3px;\n"
"  color: #fff;\n"
"  font-family: -apple-system,system-ui,\"Segoe UI\",\"Liberation Sans\",sans-serif;\n"
"  font-size: 15px;\n"
"  font-weight: 800;\n"
"  line-height: 1.15385;\n"
"  margin: 0;\n"
"  outline: none;\n"
"  padding: 8px .8em;\n"
"  text-align: center;\n"
"  text-decoration: none;\n"
"  vertical-align: baseline;\n"
"  white-space: nowrap;\n"
"}\n"
"\n"
"QPushButton:hover {\n"
"    background-color: #019ad2;\n"
"    border-color: rgb(0, 0, 0);\n"
"}\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgb(1, 140, 190);\n"
"    border-color: rgb(255, 255, 255);\n"
"}")
        self.centralwidget = QtWidgets.QWidget(parent=ImportTAG_Window)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.frame = QtWidgets.QFrame(parent=self.centralwidget)
        self.frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.frame.setObjectName("frame")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.frame)
        self.verticalLayout.setObjectName("verticalLayout")
        self.hLayout = QtWidgets.QHBoxLayout()
        self.hLayout.setObjectName("hLayout")
        self.label_SelectFile = QtWidgets.QLabel(parent=self.frame)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_SelectFile.setFont(font)
        self.label_SelectFile.setObjectName("label_SelectFile")
        self.hLayout.addWidget(self.label_SelectFile)
        self.Button_Select = QtWidgets.QPushButton(parent=self.frame)
        self.Button_Select.setMinimumSize(QtCore.QSize(250, 35))
        self.Button_Select.setMaximumSize(QtCore.QSize(250, 35))
        self.Button_Select.setObjectName("Button_Select")
        self.hLayout.addWidget(self.Button_Select)
        self.verticalLayout.addLayout(self.hLayout)
        self.label_name_file = QtWidgets.QLabel(parent=self.frame)
        self.label_name_file.setMinimumSize(QtCore.QSize(0, 25))
        self.label_name_file.setMaximumSize(QtCore.QSize(16777215, 25))
        self.label_name_file.setObjectName("label_name_file")
        self.verticalLayout.addWidget(self.label_name_file)
        self.hLayout1 = QtWidgets.QHBoxLayout()
        self.hLayout1.setObjectName("hLayout1")
        self.label_ItemType = QtWidgets.QLabel(parent=self.frame)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_ItemType.setFont(font)
        self.label_ItemType.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignLeft|QtCore.Qt.AlignmentFlag.AlignTop)
        self.label_ItemType.setObjectName("label_ItemType")
        self.hLayout1.addWidget(self.label_ItemType)
        self.vLayout = QtWidgets.QVBoxLayout()
        self.vLayout.setObjectName("vLayout")
        self.radioFlow = QtWidgets.QRadioButton(parent=self.frame)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.radioFlow.setFont(font)
        self.radioFlow.setObjectName("radioFlow")
        self.vLayout.addWidget(self.radioFlow)
        self.radioTemp = QtWidgets.QRadioButton(parent=self.frame)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.radioTemp.setFont(font)
        self.radioTemp.setObjectName("radioTemp")
        self.vLayout.addWidget(self.radioTemp)
        self.radioLevel = QtWidgets.QRadioButton(parent=self.frame)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.radioLevel.setFont(font)
        self.radioLevel.setObjectName("radioLevel")
        self.vLayout.addWidget(self.radioLevel)
        self.radioOthers = QtWidgets.QRadioButton(parent=self.frame)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.radioOthers.setFont(font)
        self.radioOthers.setObjectName("radioOthers")
        self.vLayout.addWidget(self.radioOthers)
        self.hLayout1.addLayout(self.vLayout)
        self.verticalLayout.addLayout(self.hLayout1)
        spacerItem = QtWidgets.QSpacerItem(20, 20, QtWidgets.QSizePolicy.Policy.Minimum, QtWidgets.QSizePolicy.Policy.Fixed)
        self.verticalLayout.addItem(spacerItem)
        self.hLayout2 = QtWidgets.QHBoxLayout()
        self.hLayout2.setObjectName("hLayout2")
        self.Button_Import = QtWidgets.QPushButton(parent=self.frame)
        self.Button_Import.setMinimumSize(QtCore.QSize(250, 35))
        self.Button_Import.setMaximumSize(QtCore.QSize(250, 35))
        self.Button_Import.setObjectName("Button_Import")
        self.hLayout2.addWidget(self.Button_Import)
        self.Button_Cancel = QtWidgets.QPushButton(parent=self.frame)
        self.Button_Cancel.setMinimumSize(QtCore.QSize(250, 35))
        self.Button_Cancel.setMaximumSize(QtCore.QSize(250, 35))
        self.Button_Cancel.setObjectName("Button_Cancel")
        self.hLayout2.addWidget(self.Button_Cancel)
        self.verticalLayout.addLayout(self.hLayout2)
        self.gridLayout.addWidget(self.frame, 0, 0, 1, 1)
        ImportTAG_Window.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(parent=ImportTAG_Window)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 640, 22))
        self.menubar.setObjectName("menubar")
        ImportTAG_Window.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(parent=ImportTAG_Window)
        self.statusbar.setObjectName("statusbar")
        ImportTAG_Window.setStatusBar(self.statusbar)

        self.retranslateUi(ImportTAG_Window)
        self.Button_Cancel.clicked.connect(ImportTAG_Window.close)
        self.Button_Select.clicked.connect(self.browsefiles) # type: ignore
        self.Button_Import.clicked.connect(self.importtag)

        QtCore.QMetaObject.connectSlotsByName(ImportTAG_Window)


    def retranslateUi(self, ImportTAG_Window):
        _translate = QtCore.QCoreApplication.translate
        ImportTAG_Window.setWindowTitle(_translate("ImportTAG_Window", "Importar TAG"))
        self.label_SelectFile.setText(_translate("ImportTAG_Window", "Seleccionar archivo:"))
        self.Button_Select.setText(_translate("ImportTAG_Window", "Seleccionar"))
        self.label_name_file.setText(_translate("ImportTAG_Window", ""))
        self.label_ItemType.setText(_translate("ImportTAG_Window", "Tipo de equipo:"))
        self.radioFlow.setText(_translate("ImportTAG_Window", "Caudal"))
        self.radioTemp.setText(_translate("ImportTAG_Window", "Temperatura"))
        self.radioLevel.setText(_translate("ImportTAG_Window", "Nivel"))
        self.radioOthers.setText(_translate("ImportTAG_Window", "Otros"))
        self.Button_Import.setText(_translate("ImportTAG_Window", "Importar"))
        self.Button_Cancel.setText(_translate("ImportTAG_Window", "Cancelar"))


    def browsefiles(self):
        self.fname = askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")],
                            title="Seleccionar archivo Excel")
        if self.fname:
            self.label_name_file.setText("Archivo: " + self.fname)


    def importtag(self):
        if self.label_name_file.text()=='':
            dlg = QtWidgets.QMessageBox()
            new_icon = QtGui.QIcon()
            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
            dlg.setWindowIcon(new_icon)
            dlg.setWindowTitle("ERP EIPSA")
            dlg.setText("Selecciona un archivo para importar")
            dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
            dlg.exec()
            del dlg, new_icon

        else:
            try:
                with open(self.fname, "r+") as f:
                    excel_file=self.label_name_file.text().split("Archivo: ")[1]

                    params = config()
                    conn = psycopg2.connect(**params)
                    cursor = conn.cursor()

                #Importing excel file into dataframe
                    df_table = pd.read_excel(excel_file, na_values=['N/A'], keep_default_na=False, skiprows=7, dtype={'plate_thk': str})
                    df_table = df_table.astype(str)
                    df_table.replace('nan', 'N/A', inplace=True)

                    if self.radioFlow.isChecked()==True:
                        table_name='tags_data.tags_flow'
                        seq_id='tags_flow_id_tag_flow_seq'
                        df_final = df_table.iloc[:,1:34]
                        filled_column_names = ["tag", "tag_state", "num_offer", "item_type", "line_size",
                                        "rating", "facing", "schedule", "flange_material", "flange_type",
                                        "tube_material", "tapping_num_size", "element_material", "plate_type", "plate_thk",
                                        "plate_std", "gasket_material", "bolts_nuts_material", "nace"]
                        empty_counts = []
                        for column in filled_column_names:
                            if df_table[column].eq('').sum() > 0:
                                empty_counts.append(column)

                        if len(empty_counts) > 0:
                            dlg = QtWidgets.QMessageBox()
                            new_icon = QtGui.QIcon()
                            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                            dlg.setWindowIcon(new_icon)
                            dlg.setWindowTitle("ERP EIPSA")
                            dlg.setText("Las siguientes columnas no pueden tener celdas vacías:\n" + 
                                        ', '.join(empty_counts) + "\n\n" + "El Excel no se importará")
                            dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                            dlg.exec()
                            del dlg, new_icon

                        else:
                            try:
                            # Loading Excel Template
                                self.wb = load_workbook(self.fname)    

                            # Editing sheet Import
                                sheet_name = "Import"

                                sql_query_id = f"SELECT last_value FROM pg_sequences WHERE schemaname = 'tags_data' AND sequencename = '{seq_id}'"
                                cursor.execute(sql_query_id)
                                id_tag = int(cursor.fetchone()[0]) + 1

                                for index, row in df_final.iterrows():
                                # Create a list of pairs for each column with value
                                    columns_values = [(column, row[column]) for column in df_final.columns if not pd.isnull(row[column])]

                                # Creating string for columns names and values
                                    columns = ', '.join([column for column, _ in columns_values])
                                    values = ', '.join([f"'{values.replace('.', ',')}'" if column in ['amount', 'plate_thk']
                                                        else ('NULL' if values == '' and column in ['num_order','contractual_date']
                                                        else f"'{values}'") for column, values in columns_values])

                                # Creating insertion query and executing it
                                    self.ws = self.wb[sheet_name]
                                    self.ws[f'A{index+9}'] = id_tag

                                    id_tag += 1

                                    sql_insertion = f"INSERT INTO {table_name} ({columns}) VALUES ({values})"
                                    cursor.execute(sql_insertion)

                                #Setting data validation as original excel
                                    relation_column_validation={'C':'S', 'I':'A', 'J':'B', 'K':'C', 'L':'D',
                                                                'M':'E', 'N':'F', 'O':'N', 'P':'Q', 'Q':'G',
                                                                'R':'H', 'S':'I', 'T':'J', 'U':'K', 'V':'L',
                                                                'W':'M', 'Z':'O'}
                                    for key, value in relation_column_validation.items():
                                        formula = f'=OFFSET(Validatos!${value}$1, 1, 0, COUNTA(Validatos!${value}:${value})-1, 1)'
                                        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
                                        self.ws.add_data_validation(dv)
                                        dv.add(f'{key}9:{key}2000')

                                self.wb.save(self.fname)
                            # Closing cursor and database connection
                                conn.commit()
                                cursor.close()

                                dlg = QtWidgets.QMessageBox()
                                new_icon = QtGui.QIcon()
                                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                                dlg.setWindowIcon(new_icon)
                                dlg.setWindowTitle("ERP EIPSA")
                                dlg.setText("Datos importados con éxito")
                                dlg.setIcon(QtWidgets.QMessageBox.Icon.Information)
                                dlg.exec()
                                del dlg, new_icon

                            except (Exception, psycopg2.DatabaseError) as error:
                                print(error)
                                dlg = QtWidgets.QMessageBox()
                                new_icon = QtGui.QIcon()
                                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                                dlg.setWindowIcon(new_icon)
                                dlg.setWindowTitle("ERP EIPSA")
                                dlg.setText(f"<html><body>Error con el tag <b>{row[df_final.columns[0]]}:</b><br><br>{str(error).splitlines()[1]}<br><br><b><u>NO SE REALIZARÁ LA IMPORTACIÓN<</b></u></body></html>")
                                dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                                dlg.exec()
                                del dlg, new_icon
                            finally:
                                if conn is not None:
                                    conn.close()

                    elif self.radioTemp.isChecked()==True:
                        table_name='tags_data.tags_temp'
                        seq_id='tags_temp_id_tag_temp_seq'
                        df_final = df_table.iloc[:,1:39]
                        filled_column_names = ["tag", "tag_state", "num_offer", "item_type", "tw_type",
                                        "flange_size", "flange_rating", "flange_facing", "material_tw", "root_diam",
                                        "tip_diam", "sensor_element", "sheath_stem_material", "sheath_stem_diam", "insulation",
                                        "temp_inf", "temp_sup", "nipple_ext_material", "nipple_ext_length", "head_case_material",
                                        "elec_conn_case_diam", "tt_cerblock", "material_flange_lj", "gasket_material", "puntal",
                                        "tube_t", "nace"]
                        empty_counts = []
                        for column in filled_column_names:
                            if df_table[column].eq('').sum() > 0:
                                empty_counts.append(column)

                        if len(empty_counts) > 0:
                            dlg = QtWidgets.QMessageBox()
                            new_icon = QtGui.QIcon()
                            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                            dlg.setWindowIcon(new_icon)
                            dlg.setWindowTitle("ERP EIPSA")
                            dlg.setText("Las siguientes columnas no pueden tener celdas vacías:\n" + 
                                        ', '.join(empty_counts) + "\n\n" + "El Excel no se importará")
                            dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                            dlg.exec()
                            del dlg, new_icon

                        else:
                            try:
                            # Loading Excel Template
                                self.wb = load_workbook(self.fname)    

                            # Editing sheet Import
                                sheet_name = "Import"

                                sql_query_id = f"SELECT last_value FROM pg_sequences WHERE schemaname = 'tags_data' AND sequencename = '{seq_id}'"
                                cursor.execute(sql_query_id)
                                id_tag = int(cursor.fetchone()[0]) + 1

                                for index, row in df_final.iterrows():
                                # Create a list of pairs for each column with value
                                    columns_values = [(column, row[column]) for column in df_final.columns if not pd.isnull(row[column])]

                                # Creating string for columns names and values
                                    columns = ', '.join([column for column, _ in columns_values])
                                    values = ', '.join([
                                                f"'{int(float(values))}'" if column in ['flange_rating', 'sheath_stem_diam', 'nipple_ext_length', 'temp_inf', 'temp_sup', 'root_diam', 'tip_diam'] and values.endswith('.0')
                                                else (f"'{values.replace('.', ',')}'" if column in ['amount', 'root_diam', 'tip_diam', 'sheath_stem_diam']
                                                else ('NULL' if values == 'N/A' and column in ['std_length', 'ins_length']
                                                else ('NULL' if values == '' and column in ['num_order','contractual_date']
                                                else f"'{values}'"))) for column, values in columns_values
                                                ])

                                # Creating insertion query and executing it
                                    self.ws = self.wb[sheet_name]
                                    self.ws[f'A{index+9}'] = id_tag

                                    id_tag += 1

                                    sql_insertion = f"INSERT INTO {table_name} ({columns}) VALUES ({values})"
                                    cursor.execute(sql_insertion)

                                #Setting data validation as original excel
                                    relation_column_validation={'C':'AB', 'I':'A', 'J':'B', 'K':'C', 'L':'D',
                                                                'M':'E', 'O':'F', 'R':'I', 'S':'J', 'T':'L',
                                                                'U':'M', 'V':'N', 'W':'O', 'X':'P', 'Y':'Q',
                                                                'Z':'R', 'AA':'S', 'AB':'T', 'AC':'U', 'AD':'V',
                                                                'AE':'W', 'AF':'X', 'AG':'Z', 'AH':'AA', 'AI':'K'}
                                    for key, value in relation_column_validation.items():
                                        formula = f'=OFFSET(Validatos!${value}$1, 1, 0, COUNTA(Validatos!${value}:${value})-1, 1)'
                                        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
                                        self.ws.add_data_validation(dv)
                                        dv.add(f'{key}9:{key}2000')

                                self.wb.save(self.fname)
                            # Closing cursor and database connection
                                conn.commit()
                                cursor.close()

                                dlg = QtWidgets.QMessageBox()
                                new_icon = QtGui.QIcon()
                                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                                dlg.setWindowIcon(new_icon)
                                dlg.setWindowTitle("ERP EIPSA")
                                dlg.setText("Datos importados con éxito")
                                dlg.setIcon(QtWidgets.QMessageBox.Icon.Information)
                                dlg.exec()
                                del dlg, new_icon

                            except (Exception, psycopg2.DatabaseError) as error:
                                print(error)
                                dlg = QtWidgets.QMessageBox()
                                new_icon = QtGui.QIcon()
                                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                                dlg.setWindowIcon(new_icon)
                                dlg.setWindowTitle("ERP EIPSA")
                                dlg.setText(f"<html><body>Error con el tag <b>{row[df_final.columns[0]]}:</b><br><br>{str(error).splitlines()[1]}<br><br><b><u>NO SE REALIZARÁ LA IMPORTACIÓN<</b></u></body></html>")
                                dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                                dlg.exec()
                                del dlg, new_icon
                            finally:
                                if conn is not None:
                                    conn.close()

                    elif self.radioLevel.isChecked()==True:
                        table_name='tags_data.tags_level'
                        seq_id='tags_level_id_tag_level_seq'
                        df_final = df_table.iloc[:,1:40]
                        filled_column_names = ["tag", "tag_state", "num_offer", "item_type", "model_num",
                                                "body_material", "proc_conn_type", "proc_conn_size", "proc_conn_rating", "proc_conn_facing",
                                                "conn_type", "valve_type", "dv_conn", "dv_size", "dv_rating",
                                                "dv_facing", "gasket_mica", "stud_nuts_material", "illuminator", "float_material",
                                                "case_cover_material", "scale_type", "flags", "ip_code", "flange_type",
                                                "nipple_hex", "nipple_tub", "antifrost", "nace"]
                        empty_counts = []
                        for column in filled_column_names:
                            if df_table[column].eq('').sum() > 0:
                                empty_counts.append(column)

                        if len(empty_counts) > 0:
                            dlg = QtWidgets.QMessageBox()
                            new_icon = QtGui.QIcon()
                            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                            dlg.setWindowIcon(new_icon)
                            dlg.setWindowTitle("ERP EIPSA")
                            dlg.setText("Las siguientes columnas no pueden tener celdas vacías:\n" + 
                                        ', '.join(empty_counts) + "\n\n" + "El Excel no se importará")
                            dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                            dlg.exec()
                            del dlg, new_icon

                        else:
                            try:
                            # Loading Excel Template
                                self.wb = load_workbook(self.fname)  

                            # Editing sheet Import
                                sheet_name = "Import"

                                sql_query_id = f"SELECT last_value FROM pg_sequences WHERE schemaname = 'tags_data' AND sequencename = '{seq_id}'"
                                cursor.execute(sql_query_id)
                                id_tag = int(cursor.fetchone()[0]) + 1

                                for index, row in df_final.iterrows():
                                # Create a list of pairs for each column with value
                                    columns_values = [(column, row[column]) for column in df_final.columns if not pd.isnull(row[column])]

                                # Creating string for columns names and values
                                    columns = ', '.join([column for column, _ in columns_values])
                                    values = ', '.join([
                                                f"'{int(float(values))}'" if column in ['proc_conn_rating', 'dv_rating'] and values.endswith('.0')
                                                else (f"'{values.replace('.', ',')}'" if column in ['amount']
                                                else ('NULL' if values == 'N/A' and column in ['visibility', 'cc_length']
                                                else ('NULL' if values == '' and column in ['num_order','contractual_date']
                                                else f"'{values}'"))) for column, values in columns_values
                                                ])

                                # Creating insertion query and executing it
                                    self.ws = self.wb[sheet_name]
                                    self.ws[f'A{index+9}'] = id_tag

                                    id_tag += 1

                                    sql_insertion = f"INSERT INTO {table_name} ({columns}) VALUES ({values})"
                                    cursor.execute(sql_insertion)

                                #Setting data validation as original excel
                                    relation_column_validation={'C':'AD', 'I':'A', 'J':'B', 'K':'C', 'L':'E',
                                                                'M':'F', 'N':'G', 'O':'H', 'P':'I', 'S':'J', 'T':'K',
                                                                'U':'L', 'V':'M', 'W':'N', 'X':'O', 'Y':'P',
                                                                'Z':'Q', 'AA':'R', 'AB':'S', 'AC':'T', 'AD':'U',
                                                                'AE':'V', 'AF':'W', 'AG':'X', 'AH':'X', 'AI':'Y', 'AJ':'D'}
                                    for key, value in relation_column_validation.items():
                                        formula = f'=OFFSET(Validatos!${value}$1, 1, 0, COUNTA(Validatos!${value}:${value})-1, 1)'
                                        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
                                        self.ws.add_data_validation(dv)
                                        dv.add(f'{key}9:{key}2000')

                                self.wb.save(self.fname)

                            # Closing cursor and database connection
                                conn.commit()
                                cursor.close()

                                dlg = QtWidgets.QMessageBox()
                                new_icon = QtGui.QIcon()
                                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                                dlg.setWindowIcon(new_icon)
                                dlg.setWindowTitle("ERP EIPSA")
                                dlg.setText("Datos importados con éxito")
                                dlg.setIcon(QtWidgets.QMessageBox.Icon.Information)
                                dlg.exec()
                                del dlg, new_icon

                            except (Exception, psycopg2.DatabaseError) as error:
                                print(error)
                                dlg = QtWidgets.QMessageBox()
                                new_icon = QtGui.QIcon()
                                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                                dlg.setWindowIcon(new_icon)
                                dlg.setWindowTitle("ERP EIPSA")
                                dlg.setText(f"<html><body>Error con el tag <b>{row[df_final.columns[0]]}:</b><br><br>{str(error).splitlines()[1]}<br><br><b><u>NO SE REALIZARÁ LA IMPORTACIÓN<</b></u></body></html>")
                                dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                                dlg.exec()
                                del dlg, new_icon
                            finally:
                                if conn is not None:
                                    conn.close()

                    elif self.radioOthers.isChecked()==True:
                        table_name='tags_data.tags_others'
                        seq_id='tags_others_id_tag_others_seq'
                        df_final = df_table.iloc[:,1:15]
                        filled_column_names = ["tag", "tag_state", "num_offer", "description", "nace"]
                        empty_counts = []
                        for column in filled_column_names:
                            if df_table[column].eq('').sum() > 0:
                                empty_counts.append(column)

                        if len(empty_counts) > 0:
                            dlg = QtWidgets.QMessageBox()
                            new_icon = QtGui.QIcon()
                            new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                            dlg.setWindowIcon(new_icon)
                            dlg.setWindowTitle("ERP EIPSA")
                            dlg.setText("Las siguientes columnas no pueden tener celdas vacías:\n" + 
                                        ', '.join(empty_counts) + "\n\n" + "El Excel no se importará")
                            dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                            dlg.exec()
                            del dlg, new_icon

                        else:
                            try:
                            # Loading Excel Template
                                self.wb = load_workbook(self.fname)  

                            # Editing sheet Import
                                sheet_name = "Import"

                                sql_query_id = f"SELECT last_value FROM pg_sequences WHERE schemaname = 'tags_data' AND sequencename = '{seq_id}'"
                                cursor.execute(sql_query_id)
                                id_tag = int(cursor.fetchone()[0]) + 1

                                for index, row in df_final.iterrows():
                                # Create a list of pairs for each column with value
                                    columns_values = [(column, row[column]) for column in df_final.columns if not pd.isnull(row[column])]

                                # Creating string for columns names and values
                                    columns = ', '.join([column for column, _ in columns_values])
                                    values = ', '.join([
                                                f"'{values.replace('.', ',')}'" if column in ['amount']
                                                else ('NULL' if values == '' and column in ['num_order','contractual_date']
                                                else f"'{values}'") for column, values in columns_values
                                                ])

                                # Creating insertion query and executing it
                                    self.ws = self.wb[sheet_name]
                                    self.ws[f'A{index+9}'] = id_tag

                                    id_tag += 1

                                    sql_insertion = f"INSERT INTO {table_name} ({columns}) VALUES ({values})"
                                    cursor.execute(sql_insertion)

                                #Setting data validation as original excel
                                    relation_column_validation={'C':'A', 'K':'B'}
                                    for key, value in relation_column_validation.items():
                                        formula = f'=OFFSET(Validatos!${value}$1, 1, 0, COUNTA(Validatos!${value}:${value})-1, 1)'
                                        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
                                        self.ws.add_data_validation(dv)
                                        dv.add(f'{key}9:{key}2000')

                                self.wb.save(self.fname)

                            # Closing cursor and database connection
                                conn.commit()
                                cursor.close()

                                dlg = QtWidgets.QMessageBox()
                                new_icon = QtGui.QIcon()
                                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                                dlg.setWindowIcon(new_icon)
                                dlg.setWindowTitle("ERP EIPSA")
                                dlg.setText("Datos importados con éxito")
                                dlg.setIcon(QtWidgets.QMessageBox.Icon.Information)
                                dlg.exec()
                                del dlg, new_icon

                            except (Exception, psycopg2.DatabaseError) as error:
                                print(error)
                                dlg = QtWidgets.QMessageBox()
                                new_icon = QtGui.QIcon()
                                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                                dlg.setWindowIcon(new_icon)
                                dlg.setWindowTitle("ERP EIPSA")
                                dlg.setText(f"<html><body>Error con el tag <b>{row[df_final.columns[0]]}:</b><br><br>{str(error).splitlines()[1]}<br><br><b><u>NO SE REALIZARÁ LA IMPORTACIÓN<</b></u></body></html>")
                                dlg.setIcon(QtWidgets.QMessageBox.Icon.Critical)
                                dlg.exec()
                                del dlg, new_icon
                            finally:
                                if conn is not None:
                                    conn.close()

                    else:
                        dlg = QtWidgets.QMessageBox()
                        new_icon = QtGui.QIcon()
                        new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                        dlg.setWindowIcon(new_icon)
                        dlg.setWindowTitle("ERP EIPSA")
                        dlg.setText("Selecciona un tipo de equipo")
                        dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                        dlg.exec()
                        del dlg, new_icon

                    self.label_name_file.setText("")

            except PermissionError:
                dlg = QtWidgets.QMessageBox()
                new_icon = QtGui.QIcon()
                new_icon.addPixmap(QtGui.QPixmap(os.path.abspath(os.path.join(basedir, "Resources/Iconos/icon.ico"))), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
                dlg.setWindowIcon(new_icon)
                dlg.setWindowTitle("ERP EIPSA")
                dlg.setText("El archivo Excel seleccionado está abierto\n" + 
                            "Debe estar cerrado para poder realizar la acción")
                dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
                dlg.exec()
                del dlg, new_icon



if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    ImportTAG_Window = QtWidgets.QMainWindow()
    ui = Ui_ImportTAG_Window()
    ui.setupUi(ImportTAG_Window)
    ImportTAG_Window.show()
    sys.exit(app.exec())
